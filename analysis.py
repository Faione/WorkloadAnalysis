import re
import os
import yaml
import json
import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from matplotlib.dates import DateFormatter
from datetime import datetime
from openpyxl.utils import get_column_letter

def load_json_value(data,path):
    if len(path) == 1:
        return data[path[0]]
    else:
        return load_json_value(data[path[0]],path[1:])


def pre_deal(cfg,data,data_config):
    # 需要丢弃数据
    drop_columns = []
    for column in data.columns.difference(['Time']):
        mean = data[column].mean()
        std = data[column].std()
        if mean == 0 and std ==0:
            drop_columns.append(column)
    if cfg["pre_deal"]["filter"]["metric"]["method"] == "drop":
        for column in data.columns.difference(['Time']):
            for pattern in cfg["pre_deal"]["filter"]["metric"]["drop_metric_regexs"]:
                if re.search(pattern, column):
                     drop_columns.append(column)
    elif cfg["pre_deal"]["filter"]["metric"]["method"] == "keep":
        for column in data.columns.difference(['Time']):
            keep = False
            for pattern in cfg["pre_deal"]["filter"]["metric"]["keep_metric_regexs"]:
                if re.search(pattern, column):
                     keep = True
            if keep is False:
                drop_columns.append(column)

    data = data.drop(columns=drop_columns) 
     

    # 附加数据默认值   
    for class_name,class_metrics in cfg["pre_deal"]["additions"].items():
        for metric_name, metric_default_value in class_metrics.items():
            data[metric_name] = metric_default_value

    # 时间段截取，防止数据太长，不好画图
    if cfg["pre_deal"]["filter"]["time"]["method"] == "time_range":
        data = data.loc[(data['Time'] >= cfg["pre_deal"]["filter"]["time"]["time_range"]["start_time"]) & (data['Time'] <= cfg["pre_deal"]["filter"]["time"]["time_range"]["end_time"])]
    elif cfg["pre_deal"]["filter"]["time"]["method"] == "select_apps":
        min_time = data['Time'].max() + 2000
        max_time = data['Time'].min()
        for select_app_name in cfg["pre_deal"]["filter"]["time"]["select_apps"]:
            for app_name,app_config in data_config.items():
                if "_" in app_name:
                    split_strs = app_name.split("_")
                    app_name = split_strs[0]
                    workload = split_strs[1] 
                if app_name == select_app_name and len(app_config["info_per_epoch"]) > 0:
                    epochs = app_config["info_per_epoch"]
                    # if cfg["pre_deal"]["filter"]["select_epoch"]["all"] == "true":
                    #     epochs = app_config["info_per_epoch"]
                    # else:
                    #     epochs = app_config["info_per_epoch"][:cfg["pre_deal"]["filter"]["select_epoch"]["epoch"]]
                    for split in epochs:
                        start_time = split["start_time"]
                        end_time = split["end_time"]
                        if cfg["pre_deal"]["config_time_unit"] == "s":
                            start_time = start_time * 1000
                            end_time = end_time * 1000
                        if start_time < min_time:
                            min_time = start_time
                        if end_time > max_time:
                            max_time = end_time
        data = data.loc[(data['Time'] >= min_time) & (data['Time'] <= max_time)]
    return data

def load_data(cfg,situation_config):
    data_file=situation_config["data"]
    config_file=situation_config["config"]
    
    # 加载数据
    data = pd.read_csv(data_file,index_col='Unnamed: 0')

    
    # 加载配置
    with open(config_file, 'r') as f:
        data_config = json.load(f)

    
    # 数据预处理
    data = pre_deal(cfg,data,data_config)
                    
    time_min = data['Time'].min() - 4000
    time_max = data['Time'].max() + 4000


    # 附加数据加载
    data_count = 0
    time_ranges = []
    for app_name,app_config in data_config.items():
        workload = "unknow"
        if "_" in app_name:
            split_strs = app_name.split("_")
            app_name = split_strs[0]
            workload = split_strs[1] 
        epochs = app_config["info_per_epoch"]
        # if cfg["pre_deal"]["filter"]["select_epoch"]["all"] == "true":
        #     epochs = app_config["info_per_epoch"]
        # else:
        #     epochs = app_config["info_per_epoch"][:cfg["pre_deal"]["filter"]["select_epoch"]["epoch"]]
        for split in epochs: 
            start_time = split["start_time"]
            end_time = split["end_time"]
            if cfg["pre_deal"]["config_time_unit"] == "s":
                start_time = start_time * 1000
                end_time = end_time * 1000
            if start_time >= time_min and end_time <= time_max:
                data_count = data_count + 1
                addition_name_cols = ["app","workload","count"]
                addition_value_cols = [app_name,workload,data_count]

                # 加载 qos 数据
                for qos_type,json_path in situation_config["qos"].items():
                    addition_name_cols.append(qos_type)
                    addition_value_cols.append(float(load_json_value(split,json_path.split("."))))
                
                # 加载 stress 数据
                for stress_type,json_path in situation_config["stress"].items():
                    addition_name_cols.append(stress_type)
                    addition_value_cols.append(float(load_json_value(split,json_path.split("."))))

                data.loc[(data['Time'] >= start_time) & (data['Time'] <= end_time),addition_name_cols] = addition_value_cols

                range = (datetime.fromtimestamp(start_time/1000),datetime.fromtimestamp(end_time/1000),datetime.fromtimestamp((start_time + end_time)/2000),f'app:{app_name},workload:{workload}',"red")
                time_ranges.append(range)
            else:
                print(f'数据未录入：{start_time},{time_min},{end_time},{time_max},{app_name},{workload}')

    
    data['Time'] = pd.to_datetime(data['Time'], unit='ms')
    data = data.sort_values(by='Time')

    return data,time_ranges

def draw_preview_picture(data,time_ranges=[],figlength=10,dir="",metrics=[],title="need to set title",ignore_metrics=[]): 
    if len(metrics) == 0:
        metrics = data.columns.difference(['Time'] + ignore_metrics)
    figsize=(figlength,len(metrics)*2)

    # 1. 创建子图
    fig, axs = plt.subplots(len(metrics), sharex=True, figsize=figsize)
    
    # 2. 在每个子图上绘制指标
    for i, metric_name in enumerate(metrics):
        axs[i].plot(data['Time'], data[metric_name], label=metric_name)
        axs[i].legend(loc='upper left')
    
        # 添加背景和描述
        for start, end,label_loc, label, color in time_ranges:
            axs[i].axvspan(start, end, color=color, alpha=0.2)
            axs[i].text(label_loc, axs[i].get_ylim()[0], label, 
                    horizontalalignment='center', 
                    verticalalignment='bottom')

        # 转化时间显示模式
        date_format = DateFormatter('%Y-%m-%d')  # 格式可以根据需要进行调整
        axs[i].xaxis.set_major_formatter(date_format)

        # 打印进度
        print(f"\r加载数据进度: {(i+1)/len(metrics):.2f}", end='', flush=True)
    
    # 3. 显示图形
    print(f"\n开始图片绘制,图片路径：{dir}/{title}.png")
    plt.title(title)
    plt.tight_layout()

    if not os.path.exists(dir):
        os.makedirs(dir)
        
    fig.savefig(f'{dir}/{title}.png', dpi=200)

    plt.close(fig)
    print("图片绘制结束")


def metric_list(cfg):
    for situation_name, situation_config in cfg["files"].items():
        data,time_ranges = load_data(cfg,situation_config)
        print(f'场景：{situation_name}')
        for column in data.columns:
            print(f'"{column}",')


def preview(cfg,subdir="preview",metrics=[]):
    for situation_name, situation_config in cfg["files"].items():
        data,time_ranges = load_data(cfg,situation_config)

        print(f"开始刻画场景{situation_name},指标总数{len(data.columns) - 4}")
        draw_preview_picture(
            data,
            figlength=int(len(data.index)/200),
            time_ranges=time_ranges,
            ignore_metrics=["app","workload","count"],
            title=situation_name,
            dir=f'{cfg["output_dir"]}/{subdir}',
            metrics=metrics
        )
        print(f"刻画场景{situation_name}完成")


def base_metric_compute(data,column,kind):
    if kind ==  'avg':
        value = data[column].mean()
    elif kind == 'std':
        value = data[column].std()
    elif kind == 'p99':
        value = data[column].quantile(0.99)
    else:
        value = 0
    return value

def extra_metric_to_xlsx(cfg,
                         data,
                         ignore_metrics=["count"],
                         grouped_metrics=['app','workload',"mem_stress","cpu_stress","cpu_load_stress"],
                         added_metric=[
                             'workload',
                             "mem_stress",
                             "cpu_stress",
                             "cpu_load_stress"
                             # 'p95', 'qps'
                         ],
                         filename="test"):
    metrics = []
    for metric in data.columns.difference(['Time'] + grouped_metrics + added_metric + ignore_metrics):
        for pattern in cfg["portrait"]["extra_metric_regexs"]:
            if re.search(pattern, metric):
                metrics.append(metric)

    remap = {}
    for metric_name in metrics:
        found = False
        for class_name,class_metric in cfg["portrait"]["remap"].items():
            if metric_name in class_metric.keys():
                if class_name not in remap.keys():
                    remap[class_name] = {}
                remap[class_name][metric_name] = class_metric[metric_name]
                found = True
        if not found:
            if "unknow" not in remap.keys():
                remap["unknow"] = {}
            remap["unknow"][metric_name] = metric_name

    base_metrics = cfg['portrait']['base_metric']

    grouped = data.groupby(grouped_metrics)
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # 首行
    title = "metric analysis"
    titleCols = len(metrics) * len(base_metrics) + 1 + len(added_metric)
    worksheet[f'{get_column_letter(1)}1']='application\metric'
    worksheet.merge_cells(f'{get_column_letter(2)}1:{get_column_letter(titleCols)}1')
    worksheet[f'{get_column_letter(2)}1'] = title
    
    # 2 行
    columns_index = 1
    worksheet[f'{get_column_letter(1)}2']='metric_class'
    start = columns_index + 1
    end = columns_index + len(added_metric)
    worksheet.merge_cells(f'{get_column_letter(start)}2:{get_column_letter(end)}2')
    worksheet[f'{get_column_letter(start)}2'] = "infos"
    columns_index = end
    for class_name in remap.keys():
        start = columns_index + 1
        end = columns_index + len(remap[class_name]) * len(base_metrics)
        worksheet.merge_cells(f'{get_column_letter(start)}2:{get_column_letter(end)}2')
        worksheet[f'{get_column_letter(start)}2'] = class_name
        columns_index = end
    
    columns_index = 1
    worksheet[f'{get_column_letter(1)}3']='name'
    for metric_rename in added_metric:            
        idx = columns_index + 1
        worksheet[f'{get_column_letter(idx)}3'] = metric_rename
        columns_index = columns_index +1
    for class_name,class_metrics in remap.items():
        for metric_rename in class_metrics:            
            start = columns_index + 1
            end = columns_index + len(base_metrics)
            
            worksheet.merge_cells(f'{get_column_letter(start)}3:{get_column_letter(end)}3')
            worksheet[f'{get_column_letter(start)}3'] = metric_rename
        
            columns_index = end
    
    
    columns_index = 1
    worksheet[f'{get_column_letter(1)}4']='class'
    for v in added_metric:
        idx = columns_index + 1
        worksheet[f'{get_column_letter(idx)}4'] = "avg"
        columns_index = columns_index +1
    for v in metrics:
        for cl in base_metrics:
            idx = columns_index + 1
            worksheet[f'{get_column_letter(idx)}4'] = cl
            columns_index = columns_index +1
    
    row_index = 5
    for infos,gp in grouped:
        worksheet[f'{get_column_letter(1)}{row_index}']= f'{gp["app"].value_counts().index[0]}_{gp["workload"].value_counts().index[0]}'
                
        columns_index = 1

        for v in added_metric:
            idx = columns_index + 1
            worksheet[f'{get_column_letter(idx)}{row_index}'] = gp[v].value_counts().index[0]
            columns_index = columns_index +1

        for class_name,class_metrics in remap.items():
            for metric_name in class_metrics.keys():
                for cl in base_metrics:
                    idx = columns_index + 1
                    
                    val = base_metric_compute(gp,metric_name,cl)
                    worksheet[f'{get_column_letter(idx)}{row_index}'] = val
                    columns_index = columns_index + 1
        row_index = row_index + 1
    workbook.save(f'{cfg["output_dir"]}/{filename}.xlsx')

def protrait_metric(cfg):
    for situation_name, situation_config in cfg["files"].items():
        print(f'开始生成场景{situation_name}指标，目标文件{cfg["output_dir"]}/{situation_name}.xlsx')
        data,_ = load_data(cfg,situation_config) 
        
        extra_metric_to_xlsx(cfg,data,filename=situation_name)
        print(f"指标计算完成")

def all_metrics_list(cfg):
    for situation_name, situation_config in cfg["files"].items():
        data,_ = load_data(cfg,situation_config) 
        for column in data.columns:
            print(f'"{column}",')
        break